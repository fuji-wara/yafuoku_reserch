import openpyxl
import requests
import time
from bs4 import BeautifulSoup


def main():

    """
    Excelファイルを読み込んで行ごとにURL抽出してscr_serial_number()でスクレイピングする。

    """

    wb = openpyxl.load_workbook("./20220517_1_wara3383172さん.xlsx")
    wb.create_sheet(index=1, title='完成')
    ws = wb.worksheets[0]
    wss = wb.worksheets[1]

    for i in ws["B7:O106"]:
        test_list = []
        for x in i:
            test_list.append(x.value)
        # 型番がなければ商品名で検索。         
        if test_list[6] is None:
            item_name = test_list[1].replace(' ', '+')
            item_name_url = f'https://auctions.yahoo.co.jp/search/search?p={item_name}'
            list = scr_serial_number(item_name_url, test_list)
            wss.append(list)
        else:
            serial_number = test_list[6].replace(' ', '+')
            brand = test_list[3].replace(' ', '+')
            # 型番が６文字以下ならブランド名＋型番で検索。
            if len(serial_number) >= 6:
                serial_number_url = f'https://auctions.yahoo.co.jp/search/search?p={serial_number}'
                list = scr_serial_number(serial_number_url, test_list)
                wss.append(list)
            else:
                serial_number_url = f'https://auctions.yahoo.co.jp/search/search?p={brand}+{serial_number}'               
                list = scr_serial_number(serial_number_url, test_list)
                wss.append(list)

    wb.save("./20220517_1_wara3383172さん.xlsx") 


def scr_serial_number(url, list): 
    """
    検索して商品があればtest_listに商品URLを追加する。
    """
    print(url)
    time.sleep(1)
    html = requests.get(url)
    soup = BeautifulSoup(html.text, "lxml")
    num = 0
    if search_confirmation(soup):
        item_url_list = []
        for i in soup.select("li.Product"):
            item_url = i.select_one("div.Product__image > a.Product__imageLink.js-rapid-override.js-browseHistory-add").get("href")
            item_url_list.append(item_url)
            num += 1
            if num >= 5:
                break

        if not item_url_list:
            print("一致する商品はありません")
            return list
            
        z = len(item_url_list)
        zz = 9 + z
        list[9:zz] = item_url_list

        return list

    else:
        print("一致する商品はありません")
        return list


def search_confirmation(soup):
    """
    一致する商品があればTrueなければFalseを返す。
    """
    a = soup.select_one("body > div.l-wrapper.cf > div.l-contents > div.l-contentsMain > div.l-contentsBody > div > div.Result__body > div.Products.Products--grid > div.Notice.u-marginT5.u-marginB20 > p")
    if a is None:
        return True
    else:
        return False

main()